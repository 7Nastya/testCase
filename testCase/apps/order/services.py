import io
import openpyxl
import google_auth_oauthlib.flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from testCase.apps.order.clients import CurrencyClient
from testCase.apps.order.models import Order


class AuthGoogle(object):

    def __init__(self, client_secret_filename, scopes):
        self.client_secret = client_secret_filename
        self.scopes = scopes
        self.flow = google_auth_oauthlib.flow.Flow.from_client_secrets_file(self.client_secret, self.scopes)
        self.flow.redirect_uris = 'http://localhost:8080/'
        self.creds = None

    def get_credentials(self):
        flow = InstalledAppFlow.from_client_secrets_file(self.client_secret, self.scopes)
        self.creds = flow.run_local_server(port=8080)
        return self.creds


class GoogleService(object):
    SCOPES = [
        'https://www.googleapis.com/auth/drive.readonly',
        'https://www.googleapis.com/auth/drive.metadata.readonly',
        'https://www.googleapis.com/auth/drive.file',
        'https://www.googleapis.com/auth/drive'
    ]

    SERVICE_ACCOUNT_FILE = '/home/nas/PycharmProjects/testCase/credrntials.json'

    def __init__(self):
        self.filename = 'test.xlsx'

    def download_file(self):
        credentials = AuthGoogle(client_secret_filename=self.SERVICE_ACCOUNT_FILE, scopes=self.SCOPES).get_credentials()
        service = build('drive', 'v3', credentials=credentials)
        file_id = '1IeyaDsmjTpcc6QjCH6wuFLtRbz6u8e40'
        request = service.files().get_media(fileId=file_id)
        fh = io.FileIO(self.filename, mode='wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print("Download %d%%" % int(status.progress() * 100))

    def get_actual_curs(self):
        client_currency = CurrencyClient().get_currency()
        return client_currency

    def parse_file(self):
        actual_curs = self.get_actual_curs()
        wookbook = openpyxl.load_workbook("test.xlsx")
        # Для чтения листа
        worksheet = wookbook.active
        actual_order_list_id = []
        order_exist = Order.objects.all()
        order_exist.delete()

        for i in range(1, worksheet.max_row):
            data = {}
            for col in worksheet.iter_cols(1, 4):

                value = col[i].value
                match col[i].column:
                    case (1):
                        data['number'] = int(value)
                    case (2):
                        data['number_order'] = int(value)
                    case (3):
                        price_in_rub = actual_curs * value
                        data['price_in_rub'] = price_in_rub
                        data['price_in_usd'] = value
                    case (4):
                        data['date_supply'] = value

            order = Order.objects.filter(number=data['number']).first()
            if order:
                order.price_in_rub = data['price_in_rub']
                order.save()
            else:
                order = Order.objects.create(**data)
            actual_order_list_id.append(order.id)

    def export_file(self):
        self.download_file()
        self.parse_file()
        return 0
